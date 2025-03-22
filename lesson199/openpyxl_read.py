# openpyxl - reading and modifying data in a spreadsheet
# With this library, it's possible to read and write data in specific cells,
# format cells, insert charts, create formulas, and add images or other graphical elements to your spreadsheets.
# It's useful for automating Excel tasks such as report creation, data analysis, 
# or simplifying the handling of large amounts of data.
# Required installation: pip install openpyxl
# Documentation: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Loading an existing Excel file
workbook: Workbook = load_workbook(WORKBOOK_PATH)

# Worksheet name
sheet_name = 'My Worksheet'

# Selecting the worksheet
worksheet: Worksheet = workbook[sheet_name]

from openpyxl.cell.cell import MergedCell

row: tuple[Cell | MergedCell, ...]
for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end='\t')

        # If the cell value is 'Doe', update the age to 23
        if cell.value == 'Doe':
            if cell.row is not None:
                worksheet.cell(cell.row, 2, 23)
    print()

# Directly updating a specific cell (commented example)
# worksheet['B3'].value = 14

# Saving changes to the Excel file
workbook.save(WORKBOOK_PATH)
