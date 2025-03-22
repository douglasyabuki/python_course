# openpyxl - manipulating Workbook spreadsheets
# With this library, it's possible to read and write data in specific cells,
# format cells, insert charts, create formulas, and add images or other graphical elements to your spreadsheets.
# It's useful for automating tasks involving Excel spreadsheets, such as report creation, data analysis,
# and simplifying the handling of large datasets.
# Required installation: pip install openpyxl
# Documentation: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
# worksheet: Worksheet = workbook.active

# Name for the worksheet
sheet_name = 'My Worksheet'
# Creating the worksheet
workbook.create_sheet(sheet_name, 0)
# Selecting the worksheet
worksheet: Worksheet = workbook[sheet_name]

# Removing a worksheet
workbook.remove(workbook['Sheet'])

# Creating the headers
worksheet.cell(1, 1, 'Name')
worksheet.cell(1, 2, 'Age')
worksheet.cell(1, 3, 'Grade')

students = [
    # name      age  grade
    ['John',    14,   5.5],
    ['Doe',   13,   9.7],
    ['Joe',    15,   8.8],
    ['Tho', 16,   10],
]

# Option 1: Using nested loops (commented out)
# for i, student_row in enumerate(students, start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_column)

# Option 2: Using append (recommended)
for student in students:
    worksheet.append(student)

# Saving the workbook
workbook.save(WORKBOOK_PATH)
